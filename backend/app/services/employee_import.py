import io
import json
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from ..extensions import db
from ..models import Division, Employee, ImportBatch, ImportError as ImportErrorRow
from ..models.employee import EMPLOYEE_STATUSES, ACCOMMODATION_TYPES, GENDERS
from .codes import next_code, prefix_for


TEMPLATE_COLUMNS: list[tuple[str, str]] = [
    ("employee_code", "Leave blank to auto-generate (EMP-NNNNN)"),
    ("full_name", "Required"),
    ("qid_number", "Unique if provided"),
    ("passport_number", "Unique if provided"),
    ("visa_company", ""),
    ("division_code", "Must match an existing Division code (e.g. DIV-0001)"),
    ("designation", ""),
    ("department", ""),
    ("nationality", ""),
    ("gender", f"One of: {sorted(GENDERS)}"),
    ("mobile_number", ""),
    ("joining_date", "YYYY-MM-DD"),
    ("accommodation_required", "true / false (default true)"),
    ("accommodation_type", f"One of: {sorted(ACCOMMODATION_TYPES)}"),
    ("status", f"One of: {sorted(EMPLOYEE_STATUSES)} (default active)"),
    ("emergency_contact", ""),
    ("remarks", ""),
]


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "employees"
    ws.append([c for c, _ in TEMPLATE_COLUMNS])
    ws.append([h for _, h in TEMPLATE_COLUMNS])
    ws.append([
        "", "Ahmed Al Mansoor", "12345678901", "A12345678", "PUG Trading",
        "DIV-0001", "Sales Executive", "Sales", "Indian", "male",
        "+97455551234", "2024-04-01", "true", "shared_room", "active",
        "Wife: +97455559999", "Joined April 2024",
    ])
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _norm(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _parse_bool(v: Any) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "yes", "y", "1"):
        return True
    if s in ("false", "no", "n", "0"):
        return False
    return None


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    return datetime.fromisoformat(str(v)).date()


def _row_to_payload(row: dict) -> dict:
    out: dict[str, Any] = {}
    for k in (
        "employee_code", "full_name", "qid_number", "passport_number",
        "visa_company", "division_code", "designation", "department",
        "nationality", "gender", "mobile_number", "accommodation_type",
        "status", "emergency_contact", "remarks",
    ):
        out[k] = _norm(row.get(k))
    out["joining_date"] = _parse_date(row.get("joining_date"))
    ar = _parse_bool(row.get("accommodation_required"))
    out["accommodation_required"] = True if ar is None else ar
    return out


def import_workbook(
    *, file_bytes: bytes, filename: str, actor_id: int,
) -> tuple[ImportBatch, list[Employee]]:
    """Validate every row first; if any errors are found nothing is committed."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook is empty")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c else "" for c in next(rows_iter)]
    except StopIteration:
        raise ValueError("Workbook has no header row")

    expected = {c for c, _ in TEMPLATE_COLUMNS}
    if not expected.issubset(set(header)):
        missing = expected - set(header)
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    divisions = {d.code.lower(): d for d in Division.query.all() if d.code}
    existing_qids = {e.qid_number for e in Employee.query.with_entities(Employee.qid_number).all() if e.qid_number}
    existing_passports = {e.passport_number for e in Employee.query.with_entities(Employee.passport_number).all() if e.passport_number}
    existing_codes = {e.code for e in Employee.query.with_entities(Employee.code).all()}

    batch = ImportBatch(module="employee", filename=filename, status="pending",
                        created_by=actor_id, updated_by=actor_id)
    db.session.add(batch)
    db.session.flush()

    errors: list[ImportErrorRow] = []
    parsed: list[tuple[int, dict]] = []
    seen_qids: set[str] = set()
    seen_passports: set[str] = set()
    seen_codes: set[str] = set()
    total = 0

    for idx, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all((c is None or str(c).strip() == "") for c in raw_row):
            continue
        total += 1
        row = dict(zip(header, raw_row))

        # Skip the helper description row from the template
        if str(row.get("full_name") or "").strip().lower() == "required":
            total -= 1
            continue

        row_errors: list[str] = []
        try:
            payload = _row_to_payload(row)
        except (ValueError, TypeError) as exc:
            row_errors.append(f"parse error: {exc}")
            errors.append(ImportErrorRow(batch_id=batch.id, row_number=idx,
                                         errors="; ".join(row_errors),
                                         raw_data=json.dumps({k: str(v) if v is not None else None for k, v in row.items()})))
            continue

        if not payload["full_name"]:
            row_errors.append("full_name is required")

        gender = payload.get("gender")
        if gender and gender.lower() not in GENDERS:
            row_errors.append(f"gender must be one of {sorted(GENDERS)}")
        if gender:
            payload["gender"] = gender.lower()

        acc_type = payload.get("accommodation_type")
        if acc_type and acc_type.lower() not in ACCOMMODATION_TYPES:
            row_errors.append(f"accommodation_type must be one of {sorted(ACCOMMODATION_TYPES)}")
        if acc_type:
            payload["accommodation_type"] = acc_type.lower()

        status = payload.get("status") or "active"
        if status.lower() not in EMPLOYEE_STATUSES:
            row_errors.append(f"status must be one of {sorted(EMPLOYEE_STATUSES)}")
        payload["status"] = status.lower()

        # Division resolution
        div_code = payload.pop("division_code", None)
        if div_code:
            div = divisions.get(div_code.lower())
            if not div:
                row_errors.append(f"division_code {div_code!r} not found")
            else:
                payload["division_id"] = div.id

        qid = payload.get("qid_number")
        if qid:
            if qid in existing_qids:
                row_errors.append(f"qid_number {qid} already exists")
            elif qid in seen_qids:
                row_errors.append(f"qid_number {qid} duplicated in this file")
            else:
                seen_qids.add(qid)

        passport = payload.get("passport_number")
        if passport:
            if passport in existing_passports:
                row_errors.append(f"passport_number {passport} already exists")
            elif passport in seen_passports:
                row_errors.append(f"passport_number {passport} duplicated in this file")
            else:
                seen_passports.add(passport)

        code = payload.pop("employee_code", None)
        if code:
            if code in existing_codes or code in seen_codes:
                row_errors.append(f"employee_code {code} already exists")
            else:
                seen_codes.add(code)
                payload["code"] = code

        if row_errors:
            errors.append(ImportErrorRow(
                batch_id=batch.id, row_number=idx,
                errors="; ".join(row_errors),
                raw_data=json.dumps({k: str(v) if v is not None else None for k, v in row.items()}),
            ))
            continue

        parsed.append((idx, payload))

    batch.total_rows = total

    if errors:
        for err in errors:
            db.session.add(err)
        batch.error_rows = len(errors)
        batch.success_rows = 0
        batch.status = "failed"
        db.session.commit()
        return batch, []

    created: list[Employee] = []
    for _, payload in parsed:
        if "code" not in payload:
            payload["code"] = next_code(Employee, prefix_for("employee"), width=5)
        emp = Employee(created_by=actor_id, updated_by=actor_id, **payload)
        db.session.add(emp)
        created.append(emp)

    batch.success_rows = len(created)
    batch.error_rows = 0
    batch.status = "completed"
    db.session.commit()
    return batch, created
