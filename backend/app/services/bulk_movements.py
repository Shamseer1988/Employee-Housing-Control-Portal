"""Bulk allocation / bulk transfer via Excel upload.

Each row is either an `assign` (employee -> empty bed) or a `transfer`
(employee with an active bed -> a different bed). The file is validated
end-to-end first: if ANY row fails, NOTHING is committed. Either every
row posts successfully or only the batch + error rows are saved so the
operator can fix the file and re-upload.
"""
from __future__ import annotations

import io
import json
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from ..extensions import db
from ..models import (
    AccommodationAssignment, Bed, Employee, ImportBatch, ImportError as ImportErrorRow,
)
from .assignments import AssignmentError, post_assignment
from .movements import post_transfer


TEMPLATE_COLUMNS: list[tuple[str, str]] = [
    ("mode", "assign  OR  transfer"),
    ("employee_code", "Required, e.g. EMP-00001"),
    ("bed_code", "Required, target bed (e.g. PROP-0001-F1-R101-B1)"),
    ("date", "YYYY-MM-DD (default: today)"),
    ("reason", "Free text"),
    ("remarks", ""),
]

ALLOWED_MODES = {"assign", "transfer"}


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "bulk-movements"
    ws.append([c for c, _ in TEMPLATE_COLUMNS])
    ws.append([h for _, h in TEMPLATE_COLUMNS])
    ws.append(["assign",   "EMP-00001", "PROP-0001-F1-R101-B1", "2026-01-15", "new joiner", ""])
    ws.append(["transfer", "EMP-00002", "PROP-0001-F1-R102-B2", "",           "room change", "noisy roommate"])
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _norm(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    return datetime.fromisoformat(str(v)).date()


def _raw_snapshot(row: dict) -> str:
    return json.dumps({k: str(v) if v is not None else None for k, v in row.items()})


def import_workbook(
    *, file_bytes: bytes, filename: str, actor_id: int,
) -> tuple[ImportBatch, dict]:
    """Validate every row, then either commit everything or nothing."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook is empty")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c else "" for c in next(rows_iter)]
    except StopIteration:
        raise ValueError("Workbook has no header row")

    expected = {c for c, _ in TEMPLATE_COLUMNS}
    if not expected.issubset(set(header)):
        missing = expected - set(header)
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    batch = ImportBatch(
        module="bulk_movement", filename=filename, status="pending",
        created_by=actor_id, updated_by=actor_id,
    )
    db.session.add(batch)
    db.session.flush()

    # Pre-load every employee + bed code seen in the file so each lookup is
    # O(1) instead of one query per row.
    employees_by_code: dict[str, Employee] = {}
    beds_by_code: dict[str, Bed] = {}

    rows: list[tuple[int, dict]] = []
    total = 0
    for idx, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(c is None or str(c).strip() == "" for c in raw_row):
            continue
        row = dict(zip(header, raw_row))
        # Skip the human-help row in the template
        if str(row.get("mode") or "").strip().lower() in ("assign  or  transfer", "assign or transfer"):
            continue
        total += 1
        rows.append((idx, row))

    if rows:
        codes_emp = {(_norm(r.get("employee_code")) or "").upper() for _, r in rows}
        codes_bed = {(_norm(r.get("bed_code")) or "").upper() for _, r in rows}
        for e in Employee.query.filter(Employee.code.in_(codes_emp)).all():
            employees_by_code[e.code.upper()] = e
        for b in Bed.query.filter(Bed.bed_code.in_(codes_bed)).all():
            beds_by_code[b.bed_code.upper()] = b

    errors: list[ImportErrorRow] = []
    parsed: list[tuple[int, dict]] = []
    seen_targets: dict[str, int] = {}  # bed_code -> row_number

    for idx, row in rows:
        row_errors: list[str] = []
        mode = (_norm(row.get("mode")) or "").lower()
        if mode not in ALLOWED_MODES:
            row_errors.append(f"mode must be one of {sorted(ALLOWED_MODES)}")

        emp_code = _norm(row.get("employee_code"))
        bed_code = _norm(row.get("bed_code"))
        if not emp_code:
            row_errors.append("employee_code is required")
        if not bed_code:
            row_errors.append("bed_code is required")

        employee = employees_by_code.get((emp_code or "").upper())
        bed = beds_by_code.get((bed_code or "").upper())
        if emp_code and employee is None:
            row_errors.append(f"employee_code {emp_code} not found")
        if bed_code and bed is None:
            row_errors.append(f"bed_code {bed_code} not found")

        if bed_code:
            key = bed_code.upper()
            if key in seen_targets:
                row_errors.append(f"bed_code {bed_code} already targeted on row {seen_targets[key]}")
            else:
                seen_targets[key] = idx

        try:
            when = _parse_date(row.get("date"))
        except (ValueError, TypeError):
            row_errors.append("date must be YYYY-MM-DD")
            when = None

        if not row_errors:
            try:
                if mode == "assign":
                    if employee.current_bed_id:
                        raise AssignmentError(
                            f"Employee {employee.code} already has an active bed; use mode=transfer"
                        )
                    # Dry run via the existing validator
                    from .assignments import _validate_assignment  # local import
                    _validate_assignment(employee_id=employee.id, bed_id=bed.id)
                else:  # transfer
                    if not employee.current_bed_id:
                        raise AssignmentError(
                            f"Employee {employee.code} has no active bed; use mode=assign"
                        )
                    if employee.current_bed_id == bed.id:
                        raise AssignmentError("Target bed equals current bed")
                    from .movements import _active_assignment_for, _validate_transfer  # local import
                    active = _active_assignment_for(employee)
                    _validate_transfer(employee, active, bed, _norm(row.get("reason")))
            except AssignmentError as exc:
                row_errors.append(str(exc))

        if row_errors:
            errors.append(ImportErrorRow(
                batch_id=batch.id, row_number=idx,
                errors="; ".join(row_errors), raw_data=_raw_snapshot(row),
            ))
            continue

        parsed.append((idx, {
            "mode": mode,
            "employee_id": employee.id,
            "bed_id": bed.id,
            "date": when,
            "reason": _norm(row.get("reason")),
            "remarks": _norm(row.get("remarks")),
        }))

    batch.total_rows = total

    if errors:
        for err in errors:
            db.session.add(err)
        batch.error_rows = len(errors)
        batch.success_rows = 0
        batch.status = "failed"
        db.session.commit()
        return batch, {"posted_assignments": [], "posted_transfers": []}

    posted_assignments: list[AccommodationAssignment] = []
    posted_transfers = []
    for _, p in parsed:
        if p["mode"] == "assign":
            txn = post_assignment(
                employee_id=p["employee_id"], bed_id=p["bed_id"],
                assignment_date=p["date"], reason=p["reason"], remarks=p["remarks"],
                actor_id=actor_id,
            )
            posted_assignments.append(txn)
        else:
            txn = post_transfer(
                employee_id=p["employee_id"], to_bed_id=p["bed_id"],
                transfer_date=p["date"], reason=p["reason"], remarks=p["remarks"],
                actor_id=actor_id,
            )
            posted_transfers.append(txn)

    batch.success_rows = len(parsed)
    batch.error_rows = 0
    batch.status = "completed"
    db.session.commit()
    return batch, {
        "posted_assignments": [t.transaction_number for t in posted_assignments],
        "posted_transfers": [t.transaction_number for t in posted_transfers],
    }
