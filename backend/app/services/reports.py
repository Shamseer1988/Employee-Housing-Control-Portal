"""Report registry and shared filter helpers.

Each report is a function returning ``{"columns": [...], "rows": [...], "meta": {...}}``.
The route layer turns the report into JSON or downloads it as Excel.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import func, case

from ..extensions import db
from ..models import (
    AccommodationAssignment, AccommodationCancellation, AccommodationTransfer,
    AuditLog, Bed, Division, Employee, EmployeeVacation, Floor,
    LandlordRenewal, MaintenanceRecord, Property, PropertyAgreement, Room,
)
from .reminders import agreement_bucket


Column = dict  # {"key": str, "label": str, "width": int | None}

REPORT_REGISTRY: dict[str, dict] = {}


def report(slug: str, title: str, category: str, description: str = ""):
    """Register a report builder."""

    def decorator(fn: Callable):
        REPORT_REGISTRY[slug] = {
            "slug": slug, "title": title, "category": category,
            "description": description, "builder": fn,
        }
        return fn

    return decorator


def list_reports() -> list[dict]:
    return [
        {k: v for k, v in r.items() if k != "builder"}
        for r in sorted(REPORT_REGISTRY.values(), key=lambda r: (r["category"], r["title"]))
    ]


def build_report(slug: str, filters: dict) -> dict:
    info = REPORT_REGISTRY.get(slug)
    if info is None:
        raise KeyError(f"Unknown report: {slug}")
    return info["builder"](filters)


# ---------- Shared filter parsing ----------

def parse_date(value, fallback: date | None = None) -> date | None:
    if value is None or value == "":
        return fallback
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    return datetime.fromisoformat(str(value)).date()


# ---------- Excel writer ----------

def to_workbook(title: str, columns: list[Column], rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1f3a8a")
    header_align = Alignment(horizontal="left", vertical="center")
    ws.append([c["label"] for c in columns])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r in rows:
        ws.append([_excel_value(r.get(c["key"])) for c in columns])

    for idx, c in enumerate(columns, start=1):
        width = c.get("width") or max(len(c["label"]) + 2, 12)
        ws.column_dimensions[get_column_letter(idx)].width = min(width, 50)

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _excel_value(v):
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, bool):
        return "yes" if v else "no"
    return v


# ============================================================
# Reports
# ============================================================

# 1. Property Occupancy
@report(
    "property-occupancy",
    "Property Occupancy Report",
    "Occupancy",
    "Per-property bed totals, occupancy % and a breakdown by bed status.",
)
def _r_property_occupancy(filters: dict) -> dict:
    q = (
        db.session.query(
            Property.id, Property.code, Property.name, Property.property_type,
            Property.city, Property.status,
            func.count(Bed.id).label("total"),
            func.sum(case((Bed.status == "occupied", 1), else_=0)).label("occupied"),
            func.sum(case((Bed.status == "empty", 1), else_=0)).label("empty"),
            func.sum(case((Bed.status == "reserved", 1), else_=0)).label("reserved"),
            func.sum(case((Bed.status == "maintenance", 1), else_=0)).label("maintenance"),
            func.sum(case((Bed.status == "blocked", 1), else_=0)).label("blocked"),
        )
        .outerjoin(Bed, Bed.property_id == Property.id)
        .group_by(Property.id)
    )
    if filters.get("status"):
        q = q.filter(Property.status == filters["status"])
    if filters.get("city"):
        q = q.filter(db.func.lower(Property.city) == filters["city"].lower())

    rows = []
    for r in q.all():
        total = int(r.total or 0)
        occ = int(r.occupied or 0)
        rows.append({
            "code": r.code, "name": r.name, "property_type": r.property_type,
            "city": r.city, "status": r.status,
            "total": total,
            "occupied": occ,
            "empty": int(r.empty or 0),
            "reserved": int(r.reserved or 0),
            "maintenance": int(r.maintenance or 0),
            "blocked": int(r.blocked or 0),
            "occupancy_percent": round(occ * 100 / total, 1) if total else 0.0,
        })
    rows.sort(key=lambda r: r["occupancy_percent"], reverse=True)
    return {
        "columns": [
            {"key": "code", "label": "Code", "width": 14},
            {"key": "name", "label": "Property", "width": 28},
            {"key": "property_type", "label": "Type", "width": 18},
            {"key": "city", "label": "City", "width": 16},
            {"key": "status", "label": "Status", "width": 12},
            {"key": "total", "label": "Total"},
            {"key": "occupied", "label": "Occupied"},
            {"key": "empty", "label": "Empty"},
            {"key": "reserved", "label": "Reserved"},
            {"key": "maintenance", "label": "Maintenance"},
            {"key": "blocked", "label": "Blocked"},
            {"key": "occupancy_percent", "label": "Occupancy %"},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 2. Room-wise Bed Allocation
@report(
    "room-bed-allocation",
    "Room-wise Bed Allocation",
    "Occupancy",
    "Every bed with its current employee and status.",
)
def _r_room_bed_allocation(filters: dict) -> dict:
    q = (
        db.session.query(Bed, Room, Floor, Property, Employee)
        .join(Room, Bed.room_id == Room.id)
        .join(Floor, Room.floor_id == Floor.id)
        .join(Property, Room.property_id == Property.id)
        .outerjoin(Employee, Employee.id == Bed.current_employee_id)
    )
    if filters.get("property_id"):
        q = q.filter(Property.id == int(filters["property_id"]))
    if filters.get("status"):
        q = q.filter(Bed.status == filters["status"])

    rows = []
    for bed, room, floor, prop, emp in q.order_by(Property.code, Floor.floor_number, Room.room_number, Bed.bed_number).all():
        rows.append({
            "property": prop.name,
            "property_code": prop.code,
            "floor": floor.floor_number,
            "room": room.room_number,
            "bed_code": bed.bed_code,
            "bed_type": bed.bed_type,
            "status": bed.status,
            "employee_code": emp.code if emp else None,
            "employee_name": emp.full_name if emp else None,
        })
    return {
        "columns": [
            {"key": "property_code", "label": "Property code", "width": 14},
            {"key": "property", "label": "Property", "width": 24},
            {"key": "floor", "label": "Floor", "width": 8},
            {"key": "room", "label": "Room", "width": 10},
            {"key": "bed_code", "label": "Bed code", "width": 22},
            {"key": "bed_type", "label": "Bed type", "width": 14},
            {"key": "status", "label": "Status", "width": 12},
            {"key": "employee_code", "label": "Employee code", "width": 14},
            {"key": "employee_name", "label": "Employee", "width": 26},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 3. Empty Beds
@report(
    "empty-beds",
    "Empty Bed Report",
    "Occupancy",
    "Every bed currently empty and available for assignment.",
)
def _r_empty_beds(filters: dict) -> dict:
    payload = _r_room_bed_allocation({**filters, "status": "empty"})
    payload["columns"] = [c for c in payload["columns"] if c["key"] not in ("employee_code", "employee_name", "status")]
    payload["rows"] = [{k: v for k, v in r.items() if k not in ("employee_code", "employee_name", "status")} for r in payload["rows"]]
    return payload


# 4. Property-wise Employee List
@report(
    "property-employees",
    "Property-wise Employee List",
    "Employees",
    "Employees assigned to a property, with division and bed details.",
)
def _r_property_employees(filters: dict) -> dict:
    q = Employee.query.filter(Employee.current_property_id.isnot(None))
    if filters.get("property_id"):
        q = q.filter(Employee.current_property_id == int(filters["property_id"]))
    if filters.get("division_id"):
        q = q.filter(Employee.division_id == int(filters["division_id"]))

    rows = []
    for e in q.order_by(Employee.current_property_id, Employee.full_name).all():
        rows.append({
            "code": e.code, "full_name": e.full_name,
            "qid": e.qid_number, "mobile": e.mobile_number,
            "designation": e.designation,
            "division": e.division.name if e.division else None,
            "property": e.current_property.name if e.current_property else None,
            "room": e.current_room.room_number if e.current_room else None,
            "bed_code": e.current_bed.bed_code if e.current_bed else None,
            "status": e.status,
        })
    return {
        "columns": [
            {"key": "code", "label": "Code", "width": 12},
            {"key": "full_name", "label": "Employee", "width": 26},
            {"key": "qid", "label": "QID", "width": 16},
            {"key": "mobile", "label": "Mobile", "width": 16},
            {"key": "designation", "label": "Designation", "width": 20},
            {"key": "division", "label": "Division", "width": 18},
            {"key": "property", "label": "Property", "width": 24},
            {"key": "room", "label": "Room", "width": 10},
            {"key": "bed_code", "label": "Bed code", "width": 22},
            {"key": "status", "label": "Status", "width": 14},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 5. Division-wise Accommodation
@report(
    "division-accommodation",
    "Division-wise Accommodation",
    "Employees",
    "Per-division headcounts: assigned vs not assigned vs on vacation.",
)
def _r_division(filters: dict) -> dict:
    q = (
        db.session.query(
            Division.id, Division.code, Division.name, Division.company_name,
            func.count(Employee.id).label("total"),
            func.sum(case((Employee.current_bed_id.isnot(None), 1), else_=0)).label("assigned"),
            func.sum(case((Employee.status == "on_vacation", 1), else_=0)).label("on_vacation"),
            func.sum(case(
                ((Employee.accommodation_required.is_(True)) & (Employee.current_bed_id.is_(None))
                 & (Employee.status.in_(("active", "on_vacation"))), 1),
                else_=0,
            )).label("pending"),
        )
        .outerjoin(Employee, Employee.division_id == Division.id)
        .group_by(Division.id)
        .order_by(Division.name)
    )
    rows = []
    for r in q.all():
        rows.append({
            "code": r.code, "name": r.name, "company": r.company_name,
            "total": int(r.total or 0),
            "assigned": int(r.assigned or 0),
            "on_vacation": int(r.on_vacation or 0),
            "pending": int(r.pending or 0),
        })
    return {
        "columns": [
            {"key": "code", "label": "Division code", "width": 14},
            {"key": "name", "label": "Division", "width": 24},
            {"key": "company", "label": "Company", "width": 24},
            {"key": "total", "label": "Total employees"},
            {"key": "assigned", "label": "Assigned"},
            {"key": "pending", "label": "Pending"},
            {"key": "on_vacation", "label": "On vacation"},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 6. Employee Accommodation History
@report(
    "employee-history",
    "Employee Accommodation History",
    "Employees",
    "Movement events for a single employee (assignments, transfers, cancellations, vacations).",
)
def _r_employee_history(filters: dict) -> dict:
    if not filters.get("employee_id"):
        return {"columns": [], "rows": [], "meta": {"count": 0, "note": "employee_id filter is required"}}
    employee_id = int(filters["employee_id"])
    from .movements import employee_movement_timeline
    timeline = employee_movement_timeline(employee_id)
    rows = []
    for e in timeline:
        rows.append({
            "type": e["type"],
            "transaction_number": e.get("transaction_number"),
            "date": e.get("date"),
            "bed": e.get("bed_code") or e.get("to_bed_code"),
            "from_bed": e.get("from_bed_code"),
            "reason": e.get("reason"),
            "status": e.get("status"),
            "remarks": e.get("remarks"),
        })
    return {
        "columns": [
            {"key": "type", "label": "Event", "width": 16},
            {"key": "transaction_number", "label": "Txn #", "width": 22},
            {"key": "date", "label": "Date", "width": 12},
            {"key": "bed", "label": "Bed", "width": 22},
            {"key": "from_bed", "label": "From bed", "width": 22},
            {"key": "reason", "label": "Reason", "width": 20},
            {"key": "status", "label": "Status", "width": 14},
            {"key": "remarks", "label": "Remarks", "width": 30},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "employee_id": employee_id},
    }


# 7. Agreement Expiry
@report(
    "agreement-expiry",
    "Landlord Agreement Expiry",
    "Property",
    "Active agreements with days-left and bucket (7/15/30/60/90/expired).",
)
def _r_agreement_expiry(filters: dict) -> dict:
    days = int(filters.get("within_days") or 365)
    today = date.today()
    cutoff = today + timedelta(days=days)
    q = (
        PropertyAgreement.query
        .filter(PropertyAgreement.is_active.is_(True))
        .filter(PropertyAgreement.expiry_date <= cutoff)
        .order_by(PropertyAgreement.expiry_date.asc())
    )
    rows = []
    for a in q.all():
        rows.append({
            "property_code": a.property.code if a.property else None,
            "property": a.property.name if a.property else None,
            "landlord": a.landlord.name if a.landlord else None,
            "agreement_number": a.agreement_number,
            "start_date": a.start_date.isoformat() if a.start_date else None,
            "expiry_date": a.expiry_date.isoformat() if a.expiry_date else None,
            "days_left": (a.expiry_date - today).days if a.expiry_date else None,
            "bucket": agreement_bucket(a.expiry_date, today) if a.expiry_date else None,
            "monthly_rent": float(a.monthly_rent) if a.monthly_rent is not None else None,
            "reminder_days": a.reminder_days_before_expiry,
        })
    return {
        "columns": [
            {"key": "property_code", "label": "Property code", "width": 14},
            {"key": "property", "label": "Property", "width": 24},
            {"key": "landlord", "label": "Landlord", "width": 22},
            {"key": "agreement_number", "label": "Agreement #", "width": 18},
            {"key": "start_date", "label": "Start", "width": 12},
            {"key": "expiry_date", "label": "Expiry", "width": 12},
            {"key": "days_left", "label": "Days left"},
            {"key": "bucket", "label": "Bucket", "width": 12},
            {"key": "monthly_rent", "label": "Monthly rent"},
            {"key": "reminder_days", "label": "Reminder days"},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "as_of": today.isoformat(), "within_days": days},
    }


# 8. Vacation Employee
@report(
    "vacation-employees",
    "Employees on Vacation",
    "Operations",
    "Active vacations with start, end, and bed disposition.",
)
def _r_vacations(filters: dict) -> dict:
    q = EmployeeVacation.query.filter_by(status="on_vacation")
    rows = []
    for v in q.order_by(EmployeeVacation.vacation_start_date).all():
        rows.append({
            "transaction_number": v.transaction_number,
            "employee_code": v.employee.code if v.employee else None,
            "employee": v.employee.full_name if v.employee else None,
            "bed_code": v.bed.bed_code if v.bed else None,
            "keep_bed_reserved": v.keep_bed_reserved,
            "start_date": v.vacation_start_date.isoformat() if v.vacation_start_date else None,
            "end_date": v.vacation_end_date.isoformat() if v.vacation_end_date else None,
            "remarks": v.remarks,
        })
    return {
        "columns": [
            {"key": "transaction_number", "label": "Txn #", "width": 22},
            {"key": "employee_code", "label": "Code", "width": 14},
            {"key": "employee", "label": "Employee", "width": 26},
            {"key": "bed_code", "label": "Bed", "width": 22},
            {"key": "keep_bed_reserved", "label": "Bed reserved?", "width": 14},
            {"key": "start_date", "label": "Start", "width": 12},
            {"key": "end_date", "label": "End", "width": 12},
            {"key": "remarks", "label": "Remarks", "width": 30},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 9. Maintenance
@report(
    "maintenance",
    "Maintenance Room/Bed Report",
    "Operations",
    "Active maintenance records across property / room / bed.",
)
def _r_maintenance(filters: dict) -> dict:
    q = MaintenanceRecord.query
    if filters.get("status"):
        q = q.filter_by(status=filters["status"])
    else:
        q = q.filter_by(status="in_progress")
    if filters.get("entity_type"):
        q = q.filter_by(entity_type=filters["entity_type"])

    rows = []
    for m in q.order_by(MaintenanceRecord.start_date.desc()).all():
        rows.append({
            "transaction_number": m.transaction_number,
            "entity_type": m.entity_type,
            "entity_id": m.entity_id,
            "property": m.property.name if m.property else None,
            "start_date": m.start_date.isoformat() if m.start_date else None,
            "expected_end_date": m.expected_end_date.isoformat() if m.expected_end_date else None,
            "actual_end_date": m.actual_end_date.isoformat() if m.actual_end_date else None,
            "reason": m.reason,
            "status": m.status,
        })
    return {
        "columns": [
            {"key": "transaction_number", "label": "Txn #", "width": 22},
            {"key": "entity_type", "label": "Type", "width": 12},
            {"key": "entity_id", "label": "Target ID"},
            {"key": "property", "label": "Property", "width": 24},
            {"key": "start_date", "label": "Start", "width": 12},
            {"key": "expected_end_date", "label": "Expected end", "width": 14},
            {"key": "actual_end_date", "label": "Actual end", "width": 14},
            {"key": "reason", "label": "Reason", "width": 26},
            {"key": "status", "label": "Status", "width": 12},
        ],
        "rows": rows,
        "meta": {"count": len(rows)},
    }


# 10. Monthly Movement
@report(
    "monthly-movement",
    "Monthly Accommodation Movement",
    "Operations",
    "Counts of assignments / transfers / cancellations / vacations per month.",
)
def _r_monthly_movement(filters: dict) -> dict:
    from .dashboard import monthly_movement
    months = int(filters.get("months") or 12)
    rows = monthly_movement(months=months)
    return {
        "columns": [
            {"key": "month", "label": "Month", "width": 12},
            {"key": "assignments", "label": "Assignments"},
            {"key": "transfers", "label": "Transfers"},
            {"key": "cancellations", "label": "Cancellations"},
            {"key": "vacations", "label": "Vacations"},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "months": months},
    }


# 11. Audit Trail
@report(
    "audit-trail",
    "Audit Trail",
    "Audit",
    "Recent audit log entries with module, action and entity reference.",
)
def _r_audit(filters: dict) -> dict:
    q = AuditLog.query
    if filters.get("module"):
        q = q.filter_by(module=filters["module"])
    if filters.get("action"):
        q = q.filter_by(action=filters["action"])
    if filters.get("from_date"):
        q = q.filter(AuditLog.created_at >= parse_date(filters["from_date"]))
    if filters.get("to_date"):
        cutoff = parse_date(filters["to_date"]) + timedelta(days=1)
        q = q.filter(AuditLog.created_at < cutoff)

    rows = []
    limit = min(int(filters.get("limit") or 1000), 5000)
    for r in q.order_by(AuditLog.id.desc()).limit(limit).all():
        rows.append({
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "username": r.username,
            "module": r.module,
            "action": r.action,
            "entity": f"{r.entity_type}#{r.entity_id}" if r.entity_type else "",
            "ip": r.ip_address,
            "remarks": r.remarks,
        })
    return {
        "columns": [
            {"key": "created_at", "label": "Time (UTC)", "width": 20},
            {"key": "username", "label": "User", "width": 14},
            {"key": "module", "label": "Module", "width": 14},
            {"key": "action", "label": "Action", "width": 14},
            {"key": "entity", "label": "Entity", "width": 24},
            {"key": "ip", "label": "IP", "width": 16},
            {"key": "remarks", "label": "Remarks", "width": 36},
        ],
        "rows": rows,
        "meta": {"count": len(rows), "limit": limit},
    }
