import io

from openpyxl import Workbook, load_workbook


HEADER = ["mode", "employee_code", "bed_code", "date", "reason", "remarks"]


def _seed(client, auth_headers):
    """Property with two beds and two employees, nobody assigned yet."""
    prop = client.post("/api/v1/properties", headers=auth_headers,
                       json={"name": "Bulk Prop", "property_type": "full_building"}
                       ).get_json()["data"]
    floor = client.post(f"/api/v1/properties/{prop['id']}/floors", headers=auth_headers,
                        json={"floor_number": "1"}).get_json()["data"]
    room = client.post(f"/api/v1/floors/{floor['id']}/rooms", headers=auth_headers,
                       json={"room_number": "101", "capacity": 2}).get_json()["data"]
    b1 = client.post(f"/api/v1/rooms/{room['id']}/beds", headers=auth_headers,
                     json={"bed_number": "1"}).get_json()["data"]
    b2 = client.post(f"/api/v1/rooms/{room['id']}/beds", headers=auth_headers,
                     json={"bed_number": "2"}).get_json()["data"]
    e1 = client.post("/api/v1/employees", headers=auth_headers,
                     json={"full_name": "Bulkie One", "gender": "male"}).get_json()["data"]
    e2 = client.post("/api/v1/employees", headers=auth_headers,
                     json={"full_name": "Bulkie Two", "gender": "male"}).get_json()["data"]
    return prop, [b1, b2], [e1, e2]


def _workbook(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _post(client, auth_headers, wb_bytes):
    return client.post(
        "/api/v1/bulk-movements/import",
        headers=auth_headers,
        data={"file": (io.BytesIO(wb_bytes), "bulk.xlsx")},
        content_type="multipart/form-data",
    )


def test_template_downloads(client, auth_headers):
    resp = client.get("/api/v1/bulk-movements/template", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header == HEADER


def test_bulk_assign_happy_path(client, auth_headers):
    _, beds, emps = _seed(client, auth_headers)
    wb = _workbook([
        ["assign", emps[0]["code"], beds[0]["bed_code"], "", "new joiner", ""],
        ["assign", emps[1]["code"], beds[1]["bed_code"], "", "new joiner", ""],
    ])
    resp = _post(client, auth_headers, wb)
    assert resp.status_code == 201, resp.get_data(as_text=True)
    body = resp.get_json()["data"]
    assert body["batch"]["status"] == "completed"
    assert body["batch"]["success_rows"] == 2
    assert len(body["summary"]["posted_assignments"]) == 2


def test_bulk_assign_validation_blocks_commit(client, auth_headers):
    _, beds, emps = _seed(client, auth_headers)
    wb = _workbook([
        ["assign", emps[0]["code"], beds[0]["bed_code"], "", "ok", ""],
        ["assign", "EMP-NOPE",      beds[1]["bed_code"], "", "bad", ""],   # bad emp
    ])
    resp = _post(client, auth_headers, wb)
    assert resp.status_code == 200
    body = resp.get_json()["data"]
    assert body["batch"]["status"] == "failed"
    assert body["batch"]["error_rows"] == 1
    assert body["batch"]["success_rows"] == 0
    # Nothing was committed: first bed still empty
    listed = client.get("/api/v1/assignments", headers=auth_headers).get_json()["data"]
    assert listed == []


def test_bulk_rejects_assign_when_already_assigned(client, auth_headers):
    _, beds, emps = _seed(client, auth_headers)
    # First, assign one normally
    client.post("/api/v1/assignments", headers=auth_headers,
                json={"employee_id": emps[0]["id"], "bed_id": beds[0]["id"]})
    wb = _workbook([
        ["assign", emps[0]["code"], beds[1]["bed_code"], "", "would dupe", ""],
    ])
    resp = _post(client, auth_headers, wb)
    body = resp.get_json()["data"]
    assert body["batch"]["status"] == "failed"
    assert "transfer" in body["errors"][0]["errors"]


def test_bulk_transfer_round_trip(client, auth_headers):
    _, beds, emps = _seed(client, auth_headers)
    # Pre-assign emp[0] to bed[0]
    client.post("/api/v1/assignments", headers=auth_headers,
                json={"employee_id": emps[0]["id"], "bed_id": beds[0]["id"]})

    wb = _workbook([
        ["transfer", emps[0]["code"], beds[1]["bed_code"], "", "room_change", ""],
    ])
    resp = _post(client, auth_headers, wb)
    assert resp.status_code == 201, resp.get_data(as_text=True)
    body = resp.get_json()["data"]
    assert body["batch"]["status"] == "completed"
    assert len(body["summary"]["posted_transfers"]) == 1


def test_bulk_duplicate_target_within_file(client, auth_headers):
    _, beds, emps = _seed(client, auth_headers)
    wb = _workbook([
        ["assign", emps[0]["code"], beds[0]["bed_code"], "", "", ""],
        ["assign", emps[1]["code"], beds[0]["bed_code"], "", "", ""],
    ])
    resp = _post(client, auth_headers, wb)
    body = resp.get_json()["data"]
    assert body["batch"]["status"] == "failed"
    assert any("already targeted" in e["errors"] for e in body["errors"])


def test_bulk_lists_batches(client, auth_headers):
    _, beds, emps = _seed(client, auth_headers)
    wb = _workbook([
        ["assign", emps[0]["code"], beds[0]["bed_code"], "", "", ""],
    ])
    _post(client, auth_headers, wb)
    listed = client.get("/api/v1/bulk-movements/batches", headers=auth_headers).get_json()
    assert listed["meta"]["count"] >= 1
    assert listed["data"][0]["module"] == "bulk_movement"
