import io
from datetime import date, timedelta
from openpyxl import load_workbook


def _seed(client, auth_headers):
    """Create one property + room + 2 beds, 1 division, 2 employees, post an assignment."""
    div = client.post("/api/v1/divisions", headers=auth_headers,
                      json={"name": "Retail"}).get_json()["data"]
    prop = client.post("/api/v1/properties", headers=auth_headers,
                       json={"name": "Doha B12", "property_type": "full_building", "city": "Doha"}).get_json()["data"]
    landlord = client.post("/api/v1/landlords", headers=auth_headers,
                           json={"name": "Mansoor"}).get_json()["data"]
    floor = client.post(f"/api/v1/properties/{prop['id']}/floors", headers=auth_headers,
                        json={"floor_number": "1"}).get_json()["data"]
    room = client.post(f"/api/v1/floors/{floor['id']}/rooms", headers=auth_headers,
                       json={"room_number": "101", "capacity": 2}).get_json()["data"]
    b1 = client.post(f"/api/v1/rooms/{room['id']}/beds", headers=auth_headers,
                     json={"bed_number": "1"}).get_json()["data"]
    b2 = client.post(f"/api/v1/rooms/{room['id']}/beds", headers=auth_headers,
                     json={"bed_number": "2"}).get_json()["data"]
    e1 = client.post("/api/v1/employees", headers=auth_headers,
                     json={"full_name": "Ahmed", "gender": "male", "division_id": div["id"]}).get_json()["data"]
    e2 = client.post("/api/v1/employees", headers=auth_headers,
                     json={"full_name": "Bilal", "gender": "male", "division_id": div["id"]}).get_json()["data"]
    client.post("/api/v1/assignments", headers=auth_headers,
                json={"employee_id": e1["id"], "bed_id": b1["id"]})
    today = date.today()
    client.post(f"/api/v1/properties/{prop['id']}/agreements", headers=auth_headers,
                json={"landlord_id": landlord["id"],
                      "start_date": (today - timedelta(days=200)).isoformat(),
                      "expiry_date": (today + timedelta(days=20)).isoformat(),
                      "monthly_rent": 10000})
    return {"division": div, "property": prop, "room": room, "beds": [b1, b2],
            "employees": [e1, e2], "landlord": landlord}


def test_list_reports_catalog(client, auth_headers):
    resp = client.get("/api/v1/reports", headers=auth_headers)
    assert resp.status_code == 200
    data = resp.get_json()["data"]
    slugs = {r["slug"] for r in data}
    for must_have in [
        "property-occupancy", "room-bed-allocation", "empty-beds",
        "property-employees", "division-accommodation", "employee-history",
        "agreement-expiry", "vacation-employees", "maintenance",
        "monthly-movement", "audit-trail",
    ]:
        assert must_have in slugs


def test_property_occupancy_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/property-occupancy", headers=auth_headers)
    assert resp.status_code == 200
    payload = resp.get_json()["data"]
    assert "columns" in payload and "rows" in payload
    row = next(r for r in payload["rows"] if r["code"] == s["property"]["code"])
    assert row["total"] == 2
    assert row["occupied"] == 1
    assert row["empty"] == 1
    assert row["occupancy_percent"] == 50.0


def test_room_bed_allocation_filter_by_property(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get(f"/api/v1/reports/room-bed-allocation?property_id={s['property']['id']}", headers=auth_headers)
    assert resp.status_code == 200
    payload = resp.get_json()["data"]
    assert payload["meta"]["count"] == 2
    occupied = [r for r in payload["rows"] if r["status"] == "occupied"]
    assert len(occupied) == 1
    assert occupied[0]["employee_name"] == "Ahmed"


def test_empty_beds_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/empty-beds", headers=auth_headers)
    payload = resp.get_json()["data"]
    assert payload["meta"]["count"] == 1
    assert payload["rows"][0]["bed_code"] == s["beds"][1]["bed_code"]


def test_property_employees_report(client, auth_headers):
    _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/property-employees", headers=auth_headers)
    payload = resp.get_json()["data"]
    assert payload["meta"]["count"] == 1
    assert payload["rows"][0]["full_name"] == "Ahmed"


def test_division_accommodation_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/division-accommodation", headers=auth_headers)
    payload = resp.get_json()["data"]
    row = next(r for r in payload["rows"] if r["code"] == s["division"]["code"])
    assert row["total"] == 2
    assert row["assigned"] == 1
    assert row["pending"] == 1


def test_employee_history_requires_id(client, auth_headers):
    resp = client.get("/api/v1/reports/employee-history", headers=auth_headers)
    payload = resp.get_json()["data"]
    assert payload["meta"].get("note")


def test_agreement_expiry_report(client, auth_headers):
    s = _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/agreement-expiry?within_days=60", headers=auth_headers)
    payload = resp.get_json()["data"]
    rows = payload["rows"]
    assert len(rows) >= 1
    only = next(r for r in rows if r["property_code"] == s["property"]["code"])
    assert only["bucket"] in ("30", "60", "expired", "7", "15")
    assert only["days_left"] is not None


def test_audit_trail_report(client, auth_headers):
    _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/audit-trail?module=assignment", headers=auth_headers)
    payload = resp.get_json()["data"]
    assert any(r["action"] == "post" for r in payload["rows"])


def test_excel_export_returns_xlsx(client, auth_headers):
    _seed(client, auth_headers)
    resp = client.get("/api/v1/reports/property-occupancy/export", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "Property" in header
    assert "Occupancy %" in header


def test_unknown_report_404(client, auth_headers):
    resp = client.get("/api/v1/reports/totally-fake", headers=auth_headers)
    assert resp.status_code == 404


def test_report_view_requires_permission(client):
    resp = client.get("/api/v1/reports/property-occupancy")
    assert resp.status_code == 401
